"""Parser for Excel spreadsheets (.xlsx, .xls)."""

from typing import List
from openpyxl import load_workbook
from app.parsers.base_parser import BaseParser, ParsedChunk


class XlsxParser(BaseParser):
    """Parse Excel files, extracting data from all sheets."""

    @staticmethod
    def supported_extensions() -> List[str]:
        return [".xlsx", ".xls"]

    async def parse(self, file_path: str, filename: str) -> List[ParsedChunk]:
        wb = load_workbook(file_path, data_only=True)
        chunks = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data = []

            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                if any(v.strip() for v in row_values):
                    rows_data.append(" | ".join(row_values))

            if rows_data:
                sheet_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows_data)
                chunks.append(ParsedChunk(
                    text=sheet_text,
                    metadata={
                        "source": filename,
                        "file_type": "xlsx",
                        "sheet_name": sheet_name,
                        "row_count": len(rows_data),
                    }
                ))

        return chunks
